"""
=============================================================
LendInsight -- Excel Workbook Generator
=============================================================
Generates LendInsight_Analysis.xlsx with 6 formatted sheets:
  1. README         - Guide to the workbook
  2. KPI_Summary    - Top-level business metrics
  3. Default_by_Segment  - Grade x Intent breakdown
  4. Income_and_Risk     - Income bracket analysis
  5. Portfolio_View      - Exposure and composition
  6. Raw_Data            - Full cleaned dataset
=============================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import warnings
warnings.filterwarnings("ignore")

CLEAN_PATH  = r"C:\data_analyst\LendInsight\01_data\clean\credit_risk_cleaned.csv"
OUTPUT_PATH = r"C:\data_analyst\LendInsight\04_excel\LendInsight_Analysis.xlsx"

print("Loading cleaned dataset...")
df = pd.read_csv(CLEAN_PATH)
print(f"  Rows: {len(df):,} | Cols: {len(df.columns)}")

# ─── Style Helpers ──────────────────────────────────────────
def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def std_font(size=10):
    return Font(size=size)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def pct(val, denom):
    return round(val * 100 / denom, 2) if denom else 0

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
RED         = "C0392B"
ORANGE      = "E67E22"
GREEN       = "27AE60"
GREY_HEADER = "F2F2F2"
WHITE       = "FFFFFF"

def style_header_row(ws, row, cols, fill_hex, font_color="FFFFFF", font_size=11):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = header_fill(fill_hex)
        cell.font   = bold_font(font_size, font_color)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    fill = header_fill("EAF2FB") if alt else header_fill(WHITE)
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.font   = std_font(10)
        cell.alignment = center()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1: README
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "README"
ws1.sheet_view.showGridLines = False

ws1.row_dimensions[1].height = 10
ws1["B2"] = "LendInsight: Lending Business Decision Support System"
ws1["B2"].font   = bold_font(18, DARK_BLUE)
ws1["B2"].alignment = left()

ws1["B3"] = "Excel Analysis Workbook — Version 1.0"
ws1["B3"].font = Font(size=12, color=MID_BLUE, italic=True)

ws1["B5"] = "Workbook Guide"
ws1["B5"].font = bold_font(12, DARK_BLUE)

guide = [
    ("Sheet",               "Description",                                           "Primary Stakeholder"),
    ("KPI_Summary",         "Top-level portfolio KPIs and default metrics",          "Senior Management"),
    ("Default_by_Segment",  "Default rate by loan grade and loan intent",            "Credit Risk Manager"),
    ("Income_and_Risk",     "Default analysis by income bracket and DTI",            "Loan Officer"),
    ("Portfolio_View",      "Exposure by grade, intent, and risk category",          "Portfolio Manager"),
    ("Raw_Data",            "Full cleaned dataset (32,411 rows, 15 columns)",        "Analytics Team"),
]

for i, (sheet, desc, stakeholder) in enumerate(guide, 7):
    ws1.cell(i, 2, sheet).font        = bold_font(10, DARK_BLUE) if i == 7 else std_font(10)
    ws1.cell(i, 3, desc).font         = bold_font(10, DARK_BLUE) if i == 7 else std_font(10)
    ws1.cell(i, 4, stakeholder).font  = bold_font(10, DARK_BLUE) if i == 7 else std_font(10)
    for col in [2, 3, 4]:
        ws1.cell(i, col).border = thin_border()
        ws1.cell(i, col).alignment = left()
    if i == 7:
        for col in [2, 3, 4]:
            ws1.cell(i, col).fill = header_fill(DARK_BLUE)
            ws1.cell(i, col).font = bold_font(10, WHITE)
    elif i % 2 == 0:
        for col in [2, 3, 4]:
            ws1.cell(i, col).fill = header_fill(LIGHT_BLUE)

ws1["B14"] = "Dataset: Credit Risk Dataset (Kaggle) | Source: Laotse | Rows: 32,411 | Cleaned by: LendInsight ETL Pipeline"
ws1["B14"].font = Font(size=9, italic=True, color="888888")

set_col_widths(ws1, [3, 22, 55, 28])

print("  Sheet 1/6: README done")

# ══════════════════════════════════════════════════════════════
# SHEET 2: KPI Summary
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("KPI_Summary")
ws2.sheet_view.showGridLines = False

# Title
ws2.merge_cells("B2:G2")
ws2["B2"] = "Portfolio KPI Summary — LendInsight"
ws2["B2"].font      = bold_font(16, WHITE)
ws2["B2"].fill      = header_fill(DARK_BLUE)
ws2["B2"].alignment = center()

ws2.merge_cells("B3:G3")
ws2["B3"] = "Source: lendsight_db | credit_risk_cleaned.csv | 32,411 loans"
ws2["B3"].font      = Font(size=10, italic=True, color="888888")
ws2["B3"].alignment = center()

# KPI calculations
total   = len(df)
defs    = df["default_flag"].sum()
def_rt  = round(defs * 100 / total, 2)
exp     = round(df["loan_amnt"].sum() / 1_000_000, 2)
avg_ln  = round(df["loan_amnt"].mean(), 2)
avg_rt  = round(df["loan_int_rate"].mean(), 2)
hi_risk = (df["risk_category"] == "HIGH").sum()
hi_pct  = round(hi_risk * 100 / total, 2)
def_exp = round(df[df["default_flag"]==1]["loan_amnt"].sum() / 1_000_000, 2)

kpis = [
    ("KPI",                        "Value",          "Business Meaning"),
    ("Total Loans",                f"{total:,}",      "Total loan records in portfolio"),
    ("Total Defaults",             f"{defs:,}",       "Loans that ended in default"),
    ("Overall Default Rate",       f"{def_rt}%",      "Core risk metric — 1 in 4.6 loans defaults"),
    ("Total Portfolio Exposure",   f"${exp}M",        "Total capital deployed across all loans"),
    ("Average Loan Amount",        f"${avg_ln:,.0f}", "Typical loan size"),
    ("Average Interest Rate",      f"{avg_rt}%",      "Average risk-adjusted pricing"),
    ("High-Risk Loans (Count)",    f"{hi_risk:,}",    "Loans classified as HIGH risk by segment model"),
    ("High-Risk % of Portfolio",   f"{hi_pct}%",      "Risk concentration indicator"),
    ("Defaulted Loan Exposure",    f"${def_exp}M",    "Capital at risk from defaulted loans"),
]

row = 5
for i, (kpi, val, meaning) in enumerate(kpis):
    ws2.cell(row, 2, kpi)
    ws2.cell(row, 4, val)
    ws2.cell(row, 6, meaning)
    if i == 0:
        for col in [2, 4, 6]:
            ws2.cell(row, col).fill = header_fill(MID_BLUE)
            ws2.cell(row, col).font = bold_font(11, WHITE)
            ws2.cell(row, col).alignment = center()
            ws2.cell(row, col).border = thin_border()
    else:
        alt = (i % 2 == 0)
        fill_hex = "EAF2FB" if alt else WHITE
        # Color-code the default rate
        if "Default Rate" in kpi:
            ws2.cell(row, 4).fill = header_fill("FADBD8")
            ws2.cell(row, 4).font = bold_font(11, RED)
        elif "High-Risk" in kpi:
            ws2.cell(row, 4).fill = header_fill("FDEBD0")
            ws2.cell(row, 4).font = bold_font(11, ORANGE)
        else:
            ws2.cell(row, 4).fill = header_fill(fill_hex)
            ws2.cell(row, 4).font = bold_font(11, DARK_BLUE)
        ws2.cell(row, 2).fill = header_fill(fill_hex)
        ws2.cell(row, 6).fill = header_fill(fill_hex)
        ws2.cell(row, 2).font = bold_font(10, DARK_BLUE)
        ws2.cell(row, 6).font = std_font(10)
        ws2.cell(row, 6).alignment = left()
        for col in [2, 4, 6]:
            ws2.cell(row, col).border = thin_border()
            if col != 6:
                ws2.cell(row, col).alignment = center()
    row += 1

# Risk category breakdown
row += 1
ws2.merge_cells(f"B{row}:G{row}")
ws2[f"B{row}"] = "Risk Category Breakdown"
ws2[f"B{row}"].font = bold_font(12, WHITE)
ws2[f"B{row}"].fill = header_fill(MID_BLUE)
ws2[f"B{row}"].alignment = center()

row += 1
headers = ["Risk Category", "Loan Count", "% of Portfolio", "Default Rate", "Total Exposure", "Avg Loan"]
for col, h in enumerate(headers, 2):
    ws2.cell(row, col, h).fill   = header_fill(DARK_BLUE)
    ws2.cell(row, col, h).font   = bold_font(10, WHITE)
    ws2.cell(row, col, h).alignment = center()
    ws2.cell(row, col, h).border = thin_border()

row += 1
risk_colors = {"HIGH": "FADBD8", "MEDIUM": "FDEBD0", "LOW": "D5F5E3"}
risk_font   = {"HIGH": RED,      "MEDIUM": ORANGE,   "LOW": GREEN}
for cat in ["HIGH", "MEDIUM", "LOW"]:
    sub  = df[df["risk_category"] == cat]
    vals = [
        cat,
        f"{len(sub):,}",
        f"{pct(len(sub), total)}%",
        f"{pct(sub['default_flag'].sum(), len(sub))}%",
        f"${round(sub['loan_amnt'].sum()/1_000_000, 1)}M",
        f"${round(sub['loan_amnt'].mean(), 0):,.0f}",
    ]
    for col, v in enumerate(vals, 2):
        c = ws2.cell(row, col, v)
        c.fill = header_fill(risk_colors[cat])
        c.font = bold_font(10, risk_font[cat]) if col == 2 else std_font(10)
        c.alignment = center()
        c.border = thin_border()
    row += 1

set_col_widths(ws2, [2, 28, 5, 16, 5, 30, 16])
print("  Sheet 2/6: KPI_Summary done")

# ══════════════════════════════════════════════════════════════
# SHEET 3: Default by Segment (Grade × Intent)
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Default_by_Segment")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("B2:I2")
ws3["B2"] = "Default Rate by Loan Grade and Loan Intent"
ws3["B2"].font = bold_font(14, WHITE)
ws3["B2"].fill = header_fill(DARK_BLUE)
ws3["B2"].alignment = center()

# Grade summary table
ws3["B4"] = "Section A: Default Rate by Loan Grade"
ws3["B4"].font = bold_font(12, DARK_BLUE)

grade_df = df.groupby("loan_grade").agg(
    total=("default_flag","count"),
    defaults=("default_flag","sum"),
    exposure=("loan_amnt","sum"),
    avg_rate=("loan_int_rate","mean")
).reset_index()
grade_df["default_rate"] = (grade_df["defaults"] / grade_df["total"] * 100).round(2)

gh = ["Loan Grade","Total Loans","Defaults","Default Rate","Total Exposure ($M)","Avg Interest Rate"]
row = 5
for col, h in enumerate(gh, 2):
    ws3.cell(row, col, h).fill = header_fill(MID_BLUE)
    ws3.cell(row, col, h).font = bold_font(10, WHITE)
    ws3.cell(row, col, h).alignment = center()
    ws3.cell(row, col, h).border = thin_border()

row = 6
for i, r in grade_df.iterrows():
    dr = r["default_rate"]
    color = RED if dr > 30 else (ORANGE if dr > 15 else GREEN)
    data = [r["loan_grade"], f"{r['total']:,}", f"{r['defaults']:,}",
            f"{dr}%", f"${r['exposure']/1e6:.1f}M", f"{r['avg_rate']:.2f}%"]
    for col, v in enumerate(data, 2):
        c = ws3.cell(row, col, v)
        c.fill   = header_fill("EAF2FB") if i % 2 == 0 else header_fill(WHITE)
        c.font   = bold_font(10, color) if col == 5 else std_font(10)
        c.alignment = center()
        c.border = thin_border()
    row += 1

# Intent summary table
row += 1
ws3.cell(row, 2, "Section B: Default Rate by Loan Intent").font = bold_font(12, DARK_BLUE)
row += 1

intent_df = df.groupby("loan_intent").agg(
    total=("default_flag","count"),
    defaults=("default_flag","sum"),
    avg_loan=("loan_amnt","mean")
).reset_index()
intent_df["default_rate"] = (intent_df["defaults"] / intent_df["total"] * 100).round(2)
intent_df = intent_df.sort_values("default_rate", ascending=False)

ih = ["Loan Intent","Total Loans","Defaults","Default Rate","Avg Loan Amount"]
for col, h in enumerate(ih, 2):
    ws3.cell(row, col, h).fill = header_fill(MID_BLUE)
    ws3.cell(row, col, h).font = bold_font(10, WHITE)
    ws3.cell(row, col, h).alignment = center()
    ws3.cell(row, col, h).border = thin_border()

row += 1
for i, r in enumerate(intent_df.itertuples(), 0):
    dr = r.default_rate
    color = RED if dr > 30 else (ORANGE if dr > 20 else GREEN)
    data = [r.loan_intent, f"{r.total:,}", f"{r.defaults:,}",
            f"{dr}%", f"${r.avg_loan:,.0f}"]
    for col, v in enumerate(data, 2):
        c = ws3.cell(row, col, v)
        c.fill   = header_fill("EAF2FB") if i % 2 == 0 else header_fill(WHITE)
        c.font   = bold_font(10, color) if col == 5 else std_font(10)
        c.alignment = center()
        c.border = thin_border()
    row += 1

set_col_widths(ws3, [2, 20, 14, 12, 16, 20, 20])
print("  Sheet 3/6: Default_by_Segment done")

# ══════════════════════════════════════════════════════════════
# SHEET 4: Income and Risk
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Income_and_Risk")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("B2:G2")
ws4["B2"] = "Default Analysis: Income Bracket & DTI Bracket"
ws4["B2"].font = bold_font(14, WHITE)
ws4["B2"].fill = header_fill(DARK_BLUE)
ws4["B2"].alignment = center()

def write_section(ws, start_row, title, data_df, col_headers, col_fields, color_col_idx):
    ws.cell(start_row, 2, title).font = bold_font(12, DARK_BLUE)
    start_row += 1
    for col, h in enumerate(col_headers, 2):
        c = ws.cell(start_row, col, h)
        c.fill = header_fill(MID_BLUE); c.font = bold_font(10, WHITE)
        c.alignment = center(); c.border = thin_border()
    start_row += 1
    for i, row_data in enumerate(data_df.itertuples(), 0):
        for col, field in enumerate(col_fields, 2):
            val = getattr(row_data, field)
            c = ws.cell(start_row, col, val)
            c.fill = header_fill("EAF2FB") if i % 2 == 0 else header_fill(WHITE)
            if col == color_col_idx:
                dr = float(str(val).replace("%",""))
                fc = RED if dr > 30 else (ORANGE if dr > 15 else GREEN)
                c.font = bold_font(10, fc)
            else:
                c.font = std_font(10)
            c.alignment = center(); c.border = thin_border()
        start_row += 1
    return start_row + 1

# Income bracket section
inc_df = df.groupby("income_bracket").agg(
    total=("default_flag","count"),
    defaults=("default_flag","sum"),
    avg_income=("person_income","mean")
).reset_index()
inc_df["default_rate"] = (inc_df["defaults"]/inc_df["total"]*100).round(2).astype(str) + "%"
inc_df["avg_income"]   = inc_df["avg_income"].round(0).apply(lambda x: f"${x:,.0f}")
inc_df["total"]        = inc_df["total"].apply(lambda x: f"{x:,}")
inc_df["defaults"]     = inc_df["defaults"].apply(lambda x: f"{x:,}")
order = ["Low Income","Middle Income","Upper-Middle Income","High Income"]
inc_df["_sort"] = inc_df["income_bracket"].map({v:i for i,v in enumerate(order)})
inc_df = inc_df.sort_values("_sort")

next_row = write_section(ws4, 4, "Section A: Default Rate by Income Bracket", inc_df,
    ["Income Bracket","Total Loans","Defaults","Default Rate","Avg Income"],
    ["income_bracket","total","defaults","default_rate","avg_income"], 5)

# DTI bracket section
dti_order = ["Low DTI","Moderate DTI","High DTI","Very High DTI"]
dti_df = df.groupby("dti_bracket").agg(
    total=("default_flag","count"),
    defaults=("default_flag","sum"),
    avg_dti=("loan_percent_income","mean")
).reset_index()
dti_df["default_rate"] = (dti_df["defaults"]/dti_df["total"]*100).round(2).astype(str) + "%"
dti_df["avg_dti"]      = dti_df["avg_dti"].round(3).apply(lambda x: f"{x:.3f}")
dti_df["total"]        = dti_df["total"].apply(lambda x: f"{x:,}")
dti_df["defaults"]     = dti_df["defaults"].apply(lambda x: f"{x:,}")
dti_df["_sort"]        = dti_df["dti_bracket"].map({v:i for i,v in enumerate(dti_order)})
dti_df = dti_df.sort_values("_sort")

write_section(ws4, next_row, "Section B: Default Rate by DTI Bracket", dti_df,
    ["DTI Bracket","Total Loans","Defaults","Default Rate","Avg DTI Ratio"],
    ["dti_bracket","total","defaults","default_rate","avg_dti"], 5)

set_col_widths(ws4, [2, 26, 14, 12, 16, 18])
print("  Sheet 4/6: Income_and_Risk done")

# ══════════════════════════════════════════════════════════════
# SHEET 5: Portfolio View
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Portfolio_View")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("B2:H2")
ws5["B2"] = "Portfolio Composition & Exposure Analysis"
ws5["B2"].font = bold_font(14, WHITE)
ws5["B2"].fill = header_fill(DARK_BLUE)
ws5["B2"].alignment = center()

# Exposure by grade
ws5["B4"] = "Section A: Total Exposure by Loan Grade"
ws5["B4"].font = bold_font(12, DARK_BLUE)

exp_df = df.groupby("loan_grade").agg(
    count=("loan_amnt","count"),
    total_exp=("loan_amnt","sum"),
    avg_loan=("loan_amnt","mean"),
    avg_rate=("loan_int_rate","mean"),
    defaults=("default_flag","sum")
).reset_index()
total_exp = exp_df["total_exp"].sum()
exp_df["pct_exp"] = (exp_df["total_exp"] / total_exp * 100).round(2)

eh = ["Grade","Loans","Total Exposure ($M)","% of Portfolio","Avg Loan","Avg Rate","Defaults"]
row = 5
for col, h in enumerate(eh, 2):
    ws5.cell(row, col, h).fill = header_fill(MID_BLUE)
    ws5.cell(row, col, h).font = bold_font(10, WHITE)
    ws5.cell(row, col, h).alignment = center()
    ws5.cell(row, col, h).border = thin_border()

row = 6
for i, r in enumerate(exp_df.itertuples(), 0):
    data = [r.loan_grade, f"{r.count:,}", f"${r.total_exp/1e6:.2f}M",
            f"{r.pct_exp}%", f"${r.avg_loan:,.0f}", f"{r.avg_rate:.2f}%", f"{r.defaults:,}"]
    for col, v in enumerate(data, 2):
        c = ws5.cell(row, col, v)
        c.fill = header_fill("EAF2FB") if i % 2 == 0 else header_fill(WHITE)
        c.font = std_font(10); c.alignment = center(); c.border = thin_border()
    row += 1

# Defaulters vs Non-defaulters comparison
row += 1
ws5.cell(row, 2, "Section B: Defaulters vs Non-Defaulters — Profile Comparison").font = bold_font(12, DARK_BLUE)
row += 1

comp_df = df.groupby("default_flag").agg(
    count=("loan_amnt","count"),
    avg_income=("person_income","mean"),
    avg_loan=("loan_amnt","mean"),
    avg_rate=("loan_int_rate","mean"),
    avg_dti=("loan_percent_income","mean"),
    avg_age=("person_age","mean"),
    avg_emp=("person_emp_length","mean")
).reset_index()

ch = ["Status","Count","Avg Income","Avg Loan","Avg Rate","Avg DTI","Avg Age","Avg Emp (yrs)"]
for col, h in enumerate(ch, 2):
    ws5.cell(row, col, h).fill = header_fill(MID_BLUE)
    ws5.cell(row, col, h).font = bold_font(10, WHITE)
    ws5.cell(row, col, h).alignment = center()
    ws5.cell(row, col, h).border = thin_border()
row += 1

for i, r in enumerate(comp_df.itertuples(), 0):
    label = "DEFAULTED" if r.default_flag == 1 else "REPAID"
    fill  = "FADBD8" if r.default_flag == 1 else "D5F5E3"
    font_c = RED if r.default_flag == 1 else GREEN
    data = [label, f"{r.count:,}", f"${r.avg_income:,.0f}", f"${r.avg_loan:,.0f}",
            f"{r.avg_rate:.2f}%", f"{r.avg_dti:.3f}", f"{r.avg_age:.1f}", f"{r.avg_emp:.1f}"]
    for col, v in enumerate(data, 2):
        c = ws5.cell(row, col, v)
        c.fill = header_fill(fill)
        c.font = bold_font(10, font_c) if col == 2 else std_font(10)
        c.alignment = center(); c.border = thin_border()
    row += 1

set_col_widths(ws5, [2, 10, 10, 20, 16, 12, 14, 10, 16])
print("  Sheet 5/6: Portfolio_View done")

# ══════════════════════════════════════════════════════════════
# SHEET 6: Raw Data
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Raw_Data")

headers = list(df.columns)
for col, h in enumerate(headers, 1):
    c = ws6.cell(1, col, h)
    c.fill = header_fill(DARK_BLUE)
    c.font = bold_font(10, WHITE)
    c.alignment = center()
    c.border = thin_border()

print("  Writing Raw_Data sheet (32,411 rows)...")
for i, row_data in enumerate(df.itertuples(index=False), 2):
    fill_hex = "EAF2FB" if i % 2 == 0 else WHITE
    for col, val in enumerate(row_data, 1):
        c = ws6.cell(i, col, val)
        c.fill = header_fill(fill_hex)
        c.font = std_font(9)
        c.alignment = center()

ws6.freeze_panes = "A2"
ws6.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

for col, _ in enumerate(headers, 1):
    ws6.column_dimensions[get_column_letter(col)].width = 20

print("  Sheet 6/6: Raw_Data done")

# ── Save workbook ────────────────────────────────────────────
import os
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"\nWorkbook saved -> {OUTPUT_PATH}")
print("Open in Excel to view all 6 sheets.")
